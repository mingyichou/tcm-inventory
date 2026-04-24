"""
頁面模組 — 庫存、進退貨登錄、盤點、品項管理、數據分析、叫貨出表、系統設定
v2: per-clinic 廠牌/啟用、整數數量、庫存日期欄、盤點歷史可編輯、進退貨可修改刪除
"""

import io
import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
from collections import defaultdict
from pypinyin import pinyin, Style
from database import get_supabase_client


# ══════════════════════════════════════════════
#  注音首碼搜尋（鍵盤對應版）
# ══════════════════════════════════════════════

BOPOMOFO_TO_KEY = {
    "ㄅ": "1", "ㄆ": "q", "ㄇ": "a", "ㄈ": "z",
    "ㄉ": "2", "ㄊ": "w", "ㄋ": "s", "ㄌ": "x",
    "ㄍ": "e", "ㄎ": "d", "ㄏ": "c",
    "ㄐ": "r", "ㄑ": "f", "ㄒ": "v",
    "ㄓ": "5", "ㄔ": "t", "ㄕ": "g", "ㄖ": "b",
    "ㄗ": "y", "ㄘ": "h", "ㄙ": "n",
    "ㄧ": "u", "ㄨ": "j", "ㄩ": "m",
    "ㄚ": "8", "ㄛ": "i", "ㄜ": "k", "ㄝ": ",",
    "ㄞ": "9", "ㄟ": "o", "ㄠ": "l", "ㄡ": ".",
    "ㄢ": "0", "ㄣ": "p", "ㄤ": ";", "ㄥ": "/",
    "ㄦ": "-",
}


def get_keyboard_initials(text: str) -> str:
    result = []
    for char_bpmf in pinyin(text, style=Style.BOPOMOFO):
        bpmf = char_bpmf[0]
        if bpmf and bpmf[0] in BOPOMOFO_TO_KEY:
            result.append(BOPOMOFO_TO_KEY[bpmf[0]])
    return "".join(result)


def get_bopomofo_sort_key(name: str) -> str:
    result = []
    for char_bpmf in pinyin(name, style=Style.BOPOMOFO):
        result.append(char_bpmf[0])
    return "".join(result)


def sort_products_by_bopomofo(products: list) -> list:
    return sorted(products, key=lambda p: (
        p.get("category_id", 0),
        get_bopomofo_sort_key(p["name"])
    ))


@st.cache_data(ttl=300)
def build_bopomofo_index(product_names: tuple) -> dict:
    index = {}
    for name in product_names:
        keys = get_keyboard_initials(name)
        stripped = name
        for prefix in ("特.", "特級"):
            if stripped.startswith(prefix):
                stripped = stripped[len(prefix):]
                break
        keys_stripped = get_keyboard_initials(stripped) if stripped != name else ""
        index[name] = (keys, keys_stripped)
    return index


def match_search(name: str, key_index: tuple, search_text: str) -> bool:
    search_lower = search_text.lower().strip()
    if not search_lower:
        return True
    if search_lower in name.lower():
        return True
    keys_full, keys_stripped = key_index
    if search_lower in keys_full:
        return True
    if keys_stripped and search_lower in keys_stripped:
        return True
    return False


def get_bopomofo_initial(name: str) -> str:
    """取得品項名稱第一個字的注音聲母，如 '葛根' -> 'ㄍ'"""
    for prefix in ("特.", "特級"):
        if name.startswith(prefix):
            name = name[len(prefix):]
            break
    bpmf_list = pinyin(name, style=Style.BOPOMOFO)
    if bpmf_list and bpmf_list[0]:
        first = bpmf_list[0][0]
        return first[0] if first else ""
    return ""


def short_date(d: str) -> str:
    """'2026-03-27' -> '03/27'"""
    return d[5:7] + "/" + d[8:10] if d and len(d) >= 10 else d or "-"


def style_banded(df, highlight_col=None):
    """每 5 行加粗底線分隔，可指定特定欄位加重點色"""
    def row_style(row):
        idx = row.name if isinstance(row.name, int) else 0
        border = "border-bottom: 2.5px solid #6A5ACD" if (idx + 1) % 5 == 0 else ""
        styles = [border] * len(row)
        if highlight_col and highlight_col in df.columns:
            col_idx = df.columns.get_loc(highlight_col)
            hl = "background-color: #E8E0F0; font-weight: bold; border-left: 3px solid #6A5ACD"
            styles[col_idx] = hl + ("; " + border if border else "")
        return styles
    # 數字欄位格式化為小數1位，None 顯示空白
    num_cols = df.select_dtypes(include=["number", "float", "int"]).columns.tolist()
    fmt = {c: lambda v: "" if pd.isna(v) else f"{v:.1f}" for c in num_cols}
    return df.style.apply(row_style, axis=1).format(fmt)


# ══════════════════════════════════════════════
#  共用資料載入函式
# ══════════════════════════════════════════════

@st.cache_data(ttl=30)
def load_products():
    sb = get_supabase_client()
    data = sb.table("products").select(
        "id, name, category_id, unit_id, spec_note, "
        "categories(name), units(name)"
    ).execute().data
    return sort_products_by_bopomofo(data)

@st.cache_data(ttl=30)
def load_brands():
    sb = get_supabase_client()
    return sb.table("brands").select("id, name").order("id").execute().data

@st.cache_data(ttl=30)
def load_categories():
    sb = get_supabase_client()
    return sb.table("categories").select("id, name, default_unit_id, default_spec_note").order("id").execute().data

@st.cache_data(ttl=30)
def load_units():
    sb = get_supabase_client()
    return sb.table("units").select("id, name").order("id").execute().data

def get_clinic_id(clinic_name: str) -> int | None:
    sb = get_supabase_client()
    resp = sb.table("clinics").select("id").eq("name", clinic_name).execute()
    return resp.data[0]["id"] if resp.data else None


def recalc_stock(product_id: int, clinic_id: int):
    """重算即時庫存 = 最後一次盤點數量 + 該盤點日期之後的進貨（依 tx_date）"""
    sb = get_supabase_client()
    # 找最後一次盤點
    last_log = sb.table("inventory_logs").select(
        "current_count_qty, log_date"
    ).eq("product_id", product_id).eq(
        "clinic_id", clinic_id
    ).order("log_date", desc=True).limit(1).execute().data

    if last_log:
        base_qty = float(last_log[0]["current_count_qty"])
        last_date = last_log[0]["log_date"]
    else:
        base_qty = 0
        last_date = "1900-01-01"

    # 加上最後盤點日期之後的進貨
    tx_after = sb.table("transactions").select(
        "change_qty"
    ).eq("product_id", product_id).eq(
        "clinic_id", clinic_id
    ).gt("tx_date", last_date).execute().data
    restock = sum(float(t["change_qty"]) for t in tx_after)

    new_stock = round(base_qty + restock, 1)
    sb.table("clinic_stock").update(
        {"current_stock": new_stock}
    ).eq("product_id", product_id).eq("clinic_id", clinic_id).execute()
    return new_stock


# ══════════════════════════════════════════════
#  廠牌縮寫
# ══════════════════════════════════════════════
BRAND_ABBR = {"莊松榮": "松", "港香蘭": "港", "天一": "一", "科達": "科", "順天堂": "順", "仙豐": "仙"}

def abbr_brand(name):
    return BRAND_ABBR.get(name, name) if name and name != "-" else ""


# ══════════════════════════════════════════════
#  庫存表 Excel 產生器
# ══════════════════════════════════════════════

def _build_stock_excel(df, clinic_name, h3, hr32, hc32, h2, hr21, hc21, h1):
    """產生庫存表 Excel：每分類一頁，注音分組+雙線分隔，全格線，A4 直式"""
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    thin = Side(style="thin")
    double = Side(style="double")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    data_font = Font(size=11)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # 判斷哪些期間有資料（有實際日期才展開 3 欄，否則只留 1 欄空白）
    # d3 有日期 → 盤+進+耗; 沒有 → 只留 1 欄空白
    periods = []
    if h3 in df.columns:
        has_d3 = not h3.startswith("盤(3)") and not h3.startswith("盤(-)")
        if has_d3:
            periods.append((h3, hr32, hc32, h3.replace("盤(", "").replace(")", "")))
        else:
            periods.append((h3, None, None, ""))
    if h2 in df.columns:
        has_d2 = not h2.startswith("盤(2)") and not h2.startswith("盤(-)")
        if has_d2:
            periods.append((h2, hr21, hc21, h2.replace("盤(", "").replace(")", "")))
        else:
            periods.append((h2, None, None, ""))
    if h1 in df.columns:
        has_d1 = not h1.startswith("盤(1)") and not h1.startswith("盤(-)")
        if has_d1:
            periods.append((h1, None, None, h1.replace("盤(", "").replace(")", "")))
        else:
            periods.append((h1, None, None, ""))

    # 組裝欄位：注音, 品項, 廠1, 廠2, [期間欄位...], 建議叫貨
    # 欄寬配合 A4 直式（總寬約 85 單位）
    col_specs = []  # (header_name, df_col_or_None, width)
    col_specs.append(("注音", None, 4))
    col_specs.append(("品項", "品項", 19))
    col_specs.append(("櫃", "櫃位", 4))
    col_specs.append(("廠1", "廠牌1", 4))
    col_specs.append(("廠2", "廠牌2", 4))

    for i, (h_inv, h_restock, h_consume, date_label) in enumerate(periods):
        if h_restock and h_consume:
            col_specs.append((date_label, h_inv, 8))
            col_specs.append(("進貨", h_restock, 7))
            col_specs.append(("耗用", h_consume, 7))
        else:
            col_specs.append((date_label or "", h_inv, 8))

    col_specs.append(("建議叫貨", "建議叫貨", 8))

    # 品項少的分類合併為一頁
    MERGE_CATS = {"水藥材", "高貴藥材", "非健保藥材"}

    def _write_cat_data(ws, cat_data, col_specs, start_row):
        """寫入一個分類的資料，回傳下一個可用行號"""
        row_num = start_row
        prev_initial = None
        for _, r in cat_data.iterrows():
            initial = get_bopomofo_initial(r["品項"])
            if prev_initial is not None and initial != prev_initial:
                for ci in range(1, len(col_specs) + 1):
                    cell = ws.cell(row=row_num - 1, column=ci)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=double)
            show_initial = initial if initial != prev_initial else ""
            prev_initial = initial

            for ci, (_, df_col, _) in enumerate(col_specs, 1):
                if ci == 1:
                    val = show_initial
                elif df_col == "品項":
                    val = r["品項"]
                elif df_col == "櫃位":
                    val = r.get("櫃位", "") or ""
                elif df_col == "廠牌1":
                    val = abbr_brand(r["廠牌1"])
                elif df_col == "廠牌2":
                    val = abbr_brand(r["廠牌2"])
                elif df_col and df_col in r.index:
                    v = r[df_col]
                    val = round(float(v), 1) if pd.notna(v) and v != "" else ""
                else:
                    val = ""
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.font = data_font
                cell.alignment = left_align if df_col == "品項" else center
                cell.border = border_thin
            row_num += 1
        return row_num

    def _write_header(ws, col_specs, row=1):
        for ci, (header, _, width) in enumerate(col_specs, 1):
            cell = ws.cell(row=row, column=ci, value=header)
            cell.font = header_font
            cell.alignment = center
            cell.border = border_thin
            ws.column_dimensions[get_column_letter(ci)].width = width

    def _setup_page(ws):
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"

    buf = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    cat_title_font = Font(bold=True, size=12)

    # 分組：一般分類各自一頁，合併分類集中一頁
    grouped = df.groupby("分類", sort=False)
    merge_groups = []

    for cat_name, cat_group in grouped:
        if cat_name in MERGE_CATS:
            merge_groups.append((cat_name, cat_group))
        else:
            # 獨立一頁
            ws = wb.create_sheet(title=str(cat_name)[:31])
            _setup_page(ws)
            _write_header(ws, col_specs)
            _write_cat_data(ws, cat_group.reset_index(drop=True), col_specs, 2)

    # 合併分類集中一頁
    if merge_groups:
        ws = wb.create_sheet(title="其他藥材")
        _setup_page(ws)
        _write_header(ws, col_specs)
        row_num = 2

        for idx, (cat_name, cat_group) in enumerate(merge_groups):
            # 分類標題行
            cell = ws.cell(row=row_num, column=1, value=f"【{cat_name}】")
            cell.font = cat_title_font
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=len(col_specs))
            row_num += 1

            row_num = _write_cat_data(ws, cat_group.reset_index(drop=True), col_specs, row_num)

            # 分類間空 2 行（無格線）
            if idx < len(merge_groups) - 1:
                row_num += 2

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════
#  叫貨單 Excel 產生器
# ══════════════════════════════════════════════

def _build_order_excel(order_data):
    """產生叫貨單 Excel：單一 sheet，依廠牌分區，4 欄品項+數量，全格線"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "叫貨單"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    for i in range(4):
        ws.column_dimensions[get_column_letter(i * 2 + 1)].width = 14
        ws.column_dimensions[get_column_letter(i * 2 + 2)].width = 5

    brand_font = Font(bold=True, size=11)
    header_font = Font(bold=True, size=9)
    data_font = Font(size=9)
    center = Alignment(horizontal="center", vertical="center")

    row_num = 1

    for brand_name, group in order_data.groupby("廠牌"):
        if not brand_name or brand_name == "-":
            brand_name = "未指定"

        cell = ws.cell(row=row_num, column=1, value=brand_name)
        cell.font = brand_font
        row_num += 1

        for i in range(4):
            c1 = ws.cell(row=row_num, column=i * 2 + 1, value="品項")
            c1.font = header_font
            c1.border = border_thin
            c2 = ws.cell(row=row_num, column=i * 2 + 2, value="數量")
            c2.font = header_font
            c2.border = border_thin
            c2.alignment = center
        row_num += 1

        items = list(group.itertuples(index=False))
        for chunk_start in range(0, len(items), 4):
            chunk = items[chunk_start:chunk_start + 4]
            for i, item in enumerate(chunk):
                c1 = ws.cell(row=row_num, column=i * 2 + 1, value=item.品項)
                c1.font = data_font
                c1.border = border_thin
                c2 = ws.cell(row=row_num, column=i * 2 + 2, value=round(float(item.叫貨數量), 1))
                c2.font = data_font
                c2.alignment = center
                c2.border = border_thin
            # 填滿空格的格線
            for i in range(len(chunk), 4):
                ws.cell(row=row_num, column=i * 2 + 1).border = border_thin
                ws.cell(row=row_num, column=i * 2 + 2).border = border_thin
            row_num += 1

        row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════
#  1. 庫存總覽頁面
# ══════════════════════════════════════════════

def page_stock_overview():
    st.header("📦 庫存")

    selected_clinic = st.session_state.get("selected_clinic", "澤豐")
    if selected_clinic == "合併檢視":
        st.info("請先選擇單一診所查看庫存。")
        return

    clinic_id = get_clinic_id(selected_clinic)
    if not clinic_id:
        st.error("找不到該診所")
        return

    sb = get_supabase_client()
    brands_data = load_brands()
    brand_map = {b["id"]: b["name"] for b in brands_data}
    categories = load_categories()

    col1, col2 = st.columns([2, 3])
    with col1:
        cat_filter = st.selectbox("分類", ["全部"] + [c["name"] for c in categories], key="stock_cat")
    with col2:
        search = st.text_input("搜尋", placeholder="中文 或 注音首碼（ee=葛根, 2e=當歸）", key="stock_search")

    # 載入品項
    products = sb.table("products").select(
        "id, name, category_id, categories(name), units(name)"
    ).execute().data
    products = sort_products_by_bopomofo(products)

    # 載入 per-clinic 資料
    cs_data = sb.table("clinic_stock").select(
        "product_id, current_stock, brand1_id, brand2_id, is_active, cabinet"
    ).eq("clinic_id", clinic_id).execute().data
    cs_map = {s["product_id"]: s for s in cs_data}

    # 搜尋索引
    all_names = tuple(p["name"] for p in products)
    key_index = build_bopomofo_index(all_names)

    # 載入系統設定
    settings = sb.table("system_settings").select("*").execute().data
    safety_factor = float(next(s["value"] for s in settings if s["key"] == "safety_factor"))
    stock_multiplier = float(next(s["value"] for s in settings if s["key"] == "stock_target_multiplier"))

    # 載入盤點紀錄
    all_logs = sb.table("inventory_logs").select(
        "product_id, current_count_qty, consumed_qty, log_date, session_id"
    ).eq("clinic_id", clinic_id).order("log_date", desc=True).execute().data

    # 計算每品項平均耗用
    consumed_by_product = defaultdict(list)
    for log in all_logs:
        if log["consumed_qty"] is not None and log["consumed_qty"] > 0:
            consumed_by_product[log["product_id"]].append(float(log["consumed_qty"]))

    # 取得最近 3 個不重複的盤點日期
    seen_dates = []
    for log in all_logs:
        d = log["log_date"]
        if d not in seen_dates:
            seen_dates.append(d)
        if len(seen_dates) >= 3:
            break
    # seen_dates: [最新, ..., 最舊]; display: 最舊→最新
    display_dates = list(reversed(seen_dates))

    # 填滿 3 個位置: d3(最舊), d2, d1(最新)
    d3 = display_dates[0] if len(display_dates) >= 3 else None
    d2 = display_dates[1] if len(display_dates) >= 3 else (display_dates[0] if len(display_dates) >= 2 else None)
    d1 = display_dates[-1] if display_dates else None

    sd3 = short_date(d3) if d3 else "-"
    sd2 = short_date(d2) if d2 else "-"
    sd1 = short_date(d1) if d1 else "-"

    # 欄位標題
    h3 = f"盤({sd3})" if d3 else "盤(3)"
    h2 = f"盤({sd2})" if d2 else "盤(2)"
    h1 = f"盤({sd1})" if d1 else "盤(1)"
    hr32 = f"進({sd3}~{sd2})" if d3 and d2 else "進(3~2)"
    hc32 = f"耗({sd3}~{sd2})" if d3 and d2 else "耗(3~2)"
    hr21 = f"進({sd2}~{sd1})" if d2 and d1 and d2 != d1 else "進(2~1)"
    hc21 = f"耗({sd2}~{sd1})" if d2 and d1 and d2 != d1 else "耗(2~1)"

    # 每品項盤點紀錄
    product_log_map = defaultdict(dict)
    for log in all_logs:
        pid = log["product_id"]
        d = log["log_date"]
        if d not in product_log_map[pid]:
            product_log_map[pid][d] = log

    # 進退貨
    all_tx = sb.table("transactions").select(
        "product_id, change_qty, tx_date"
    ).eq("clinic_id", clinic_id).order("tx_date").execute().data
    tx_by_product = defaultdict(list)
    for tx in all_tx:
        tx_by_product[tx["product_id"]].append(tx)

    # 組裝表格（固定 15 欄）
    rows = []
    product_id_list = []

    for p in products:
        pid = p["id"]
        cs = cs_map.get(pid)
        if not cs or not cs.get("is_active", True):
            continue
        if cat_filter != "全部" and p["categories"]["name"] != cat_filter:
            continue
        if search and not match_search(p["name"], key_index.get(p["name"], ("", "")), search):
            continue

        p_logs = product_log_map.get(pid, {})
        log3 = p_logs.get(d3) if d3 else None
        log2 = p_logs.get(d2) if d2 else None
        log1 = p_logs.get(d1) if d1 else None

        v3 = round(float(log3["current_count_qty"]), 1) if log3 else None
        v2 = round(float(log2["current_count_qty"]), 1) if log2 else None
        v1 = round(float(log1["current_count_qty"]), 1) if log1 else None

        # 進貨 3→2
        if d3 and d2 and d3 != d2:
            r32 = sum(float(t["change_qty"]) for t in tx_by_product.get(pid, []) if d3 < t["tx_date"] <= d2)
        else:
            r32 = None
        c32 = (v3 + r32 - v2) if (v3 is not None and v2 is not None and r32 is not None) else None

        # 進貨 2→1
        if d2 and d1 and d2 != d1:
            r21 = sum(float(t["change_qty"]) for t in tx_by_product.get(pid, []) if d2 < t["tx_date"] <= d1)
        else:
            r21 = None
        c21 = (v2 + r21 - v1) if (v2 is not None and v1 is not None and r21 is not None) else None

        # 進貨迄今
        r_now = sum(float(t["change_qty"]) for t in tx_by_product.get(pid, []) if t["tx_date"] > d1) if d1 else 0

        current_stock = float(cs["current_stock"]) if cs else 0

        # 建議叫貨
        avg_vals = consumed_by_product.get(pid, [])
        avg_c = sum(avg_vals) / len(avg_vals) if avg_vals else 0
        if avg_c > 0 and current_stock <= avg_c * safety_factor:
            suggested = max(0, round(avg_c * stock_multiplier - current_stock, 1))
        else:
            suggested = 0

        row = {
            "品項": p["name"],
            "分類": p["categories"]["name"],
            "櫃位": cs.get("cabinet") or "",
            "廠牌1": brand_map.get(cs.get("brand1_id"), "-"),
            "廠牌2": brand_map.get(cs.get("brand2_id"), "-"),
            h3: v3, hr32: r32, hc32: c32,
            h2: v2, hr21: r21, hc21: c21,
            h1: v1,
            "進(迄今)": r_now,
            "即時庫存": current_stock,
            "建議叫貨": suggested,
            "叫貨": suggested,
        }
        rows.append(row)
        product_id_list.append(pid)

    if not rows:
        st.info("沒有符合條件的品項")
        return

    df = pd.DataFrame(rows)

    # 主表格（data_editor：叫貨欄可編輯，其餘唯讀）
    col_config = {
        "品項": st.column_config.TextColumn(disabled=True),
        "分類": st.column_config.TextColumn(disabled=True),
        "廠牌1": st.column_config.TextColumn(disabled=True),
        "櫃位": st.column_config.TextColumn(disabled=True),
        "廠牌2": st.column_config.TextColumn(disabled=True),
        h3: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        hr32: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        hc32: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        h2: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        hr21: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        hc21: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        h1: st.column_config.NumberColumn(disabled=True, format="%.1f"),
        "進(迄今)": st.column_config.NumberColumn(disabled=True, format="%.1f"),
        "即時庫存": st.column_config.NumberColumn(disabled=True, format="%.1f"),
        "建議叫貨": st.column_config.NumberColumn(disabled=True, format="%.1f"),
        "叫貨": st.column_config.NumberColumn("叫貨 ✏️", min_value=0, format="%.1f"),
    }

    edited_stock_df = st.data_editor(
        df, use_container_width=True, hide_index=True,
        height=min(len(df) * 35 + 38, 700),
        column_config=col_config, key="stock_editor",
    )

    # 送出叫貨
    order_items = edited_stock_df[edited_stock_df["叫貨"] > 0]
    if not order_items.empty:
        st.markdown(f"**共 {len(order_items)} 項需要叫貨**")
        if st.button("🛒 送出叫貨清單", type="primary"):
            order_data = order_items[["品項", "分類", "廠牌1", "即時庫存", "叫貨"]].copy()
            order_data.columns = ["品項", "分類", "廠牌", "目前庫存", "叫貨數量"]

            st.subheader("叫貨清單")
            for bname, group in order_data.groupby("廠牌"):
                if not bname or bname == "-":
                    bname = "未指定"
                st.markdown(f"### 【{bname}】")
                st.dataframe(group[["品項", "叫貨數量"]], use_container_width=True, hide_index=True)

            buf = _build_order_excel(order_data)
            st.download_button("📥 下載叫貨單 (.xlsx)", data=buf,
                file_name=f"叫貨單_{selected_clinic}_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    # 匯出庫存表
    st.divider()
    if st.button("📥 匯出庫存表 (.xlsx)", use_container_width=True):
        buf = _build_stock_excel(df, selected_clinic, h3, hr32, hc32, h2, hr21, hc21, h1)
        st.download_button(
            "📥 下載庫存表",
            data=buf,
            file_name=f"庫存表_{selected_clinic}_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
        )

    # 改錯存檔
    with st.expander("🔧 改錯存檔（修正盤點數量）"):
        st.caption("僅供事後檢討發現錯誤時修正盤點數量")
        if not product_id_list:
            return

        product_names = {rows[i]["品項"]: product_id_list[i] for i in range(len(rows))}
        selected_name = st.selectbox("選擇品項", list(product_names.keys()), key="fix_product")
        fix_pid = product_names[selected_name]

        fix_logs = sb.table("inventory_logs").select(
            "id, current_count_qty, log_date, session_id"
        ).eq("product_id", fix_pid).eq("clinic_id", clinic_id).order("log_date", desc=True).limit(5).execute().data

        if not fix_logs:
            st.info("該品項尚無盤點紀錄")
        else:
            fix_options = {f"{l['log_date']} — 數量: {round(float(l['current_count_qty']), 1)}": l for l in fix_logs}
            selected_fix = st.selectbox("選擇要修正的盤點紀錄", list(fix_options.keys()), key="fix_log")
            fix_log = fix_options[selected_fix]

            new_qty = st.number_input("修正後數量", value=float(fix_log["current_count_qty"]), step=0.1, format="%.1f", key="fix_qty")

            if st.button("改錯存檔", type="primary"):
                try:
                    sb.table("inventory_logs").update(
                        {"current_count_qty": new_qty}
                    ).eq("id", fix_log["id"]).execute()

                    recalc_stock(fix_pid, clinic_id)
                    st.success(f"已修正：{selected_name} {fix_log['log_date']} → {new_qty}")
                except Exception as e:
                    st.error(f"修正失敗：{e}")

    st.caption(f"共 {len(df)} 個品項 — {selected_clinic}")


# ══════════════════════════════════════════════
#  2. 進退貨登錄頁面
# ══════════════════════════════════════════════

def page_transactions():
    st.header("📥 進退貨登錄")

    selected_clinic = st.session_state.get("selected_clinic", "澤豐")
    user = st.session_state.user

    if selected_clinic == "合併檢視":
        st.info("請先選擇單一診所。")
        return

    clinic_id = get_clinic_id(selected_clinic)
    if not clinic_id:
        st.error("找不到該診所")
        return

    tab_add, tab_history = st.tabs(["➕ 新增紀錄", "📜 歷史紀錄"])

    with tab_add:
        sb = get_supabase_client()
        products = sb.table("products").select(
            "id, name, category_id, categories(name), units(name)"
        ).execute().data
        products = sort_products_by_bopomofo(products)

        # per-clinic 啟用篩選
        cs_data = sb.table("clinic_stock").select(
            "product_id, current_stock, is_active"
        ).eq("clinic_id", clinic_id).execute().data
        cs_map = {s["product_id"]: s for s in cs_data}
        products = [p for p in products if cs_map.get(p["id"], {}).get("is_active", True)]

        stock_map = {s["product_id"]: float(s["current_stock"]) for s in cs_data}

        categories = load_categories()
        col_filter, col_search = st.columns([1, 2])
        with col_filter:
            cat_filter = st.selectbox("分類", ["全部"] + [c["name"] for c in categories], key="tx_cat")
        with col_search:
            search = st.text_input("搜尋品項", placeholder="中文 或 注音首碼", key="tx_search")

        tx_names = tuple(p["name"] for p in products)
        tx_key_index = build_bopomofo_index(tx_names)

        filtered = products
        if cat_filter != "全部":
            filtered = [p for p in filtered if p["categories"]["name"] == cat_filter]
        if search:
            filtered = [p for p in filtered if match_search(p["name"], tx_key_index.get(p["name"], ("", "")), search)]

        if not filtered:
            st.warning("沒有符合條件的品項")
            return

        product_options = {f"{p['name']}（{p['categories']['name']}）": p for p in filtered}
        selected = st.selectbox("選擇品項", list(product_options.keys()), key="tx_product")
        product = product_options[selected]

        current_stock = stock_map.get(product["id"], 0)
        st.metric(f"📍 {selected_clinic} 目前庫存", f"{current_stock} {product['units']['name']}")

        with st.form("tx_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                tx_type = st.selectbox("異動類型", ["進貨", "調撥", "廢棄"])
            with col2:
                qty = st.number_input("數量（負數=調出/廢棄）", value=0.0, step=0.1, format="%.1f")
            with col3:
                tx_date = st.date_input("日期", value=date.today())

            col4, col5 = st.columns(2)
            with col4:
                tx_brands = load_brands()
                tx_brand_options = ["-"] + [b["name"] for b in tx_brands]
                tx_brand = st.selectbox("廠牌", tx_brand_options, key="tx_brand")
            with col5:
                note = st.text_input("備註（選填）")

            if st.form_submit_button("確認登錄", type="primary", use_container_width=True):
                if qty == 0:
                    st.error("數量不可為 0")
                else:
                    try:
                        full_note = f"[{tx_brand}] {note}".strip() if tx_brand != "-" else (note or None)
                        sb.table("transactions").insert({
                            "product_id": product["id"], "clinic_id": clinic_id,
                            "change_qty": qty, "tx_date": str(tx_date),
                            "tx_type": tx_type, "note": full_note, "created_by": user["id"],
                        }).execute()

                        new_stock = recalc_stock(product["id"], clinic_id)
                        st.success(f"已登錄：{product['name']} {qty:+.1f}（即時庫存：{new_stock}）")
                    except Exception as e:
                        st.error(f"登錄失敗：{e}")

    with tab_history:
        sb = get_supabase_client()
        col1, col2 = st.columns(2)
        with col1:
            days_filter = st.selectbox("時間範圍", ["最近 30 天", "最近 7 天", "最近 90 天", "全部"], key="tx_days")
        with col2:
            type_filter = st.selectbox("類型", ["全部", "進貨", "調撥", "廢棄"], key="tx_type_filter")

        query = sb.table("transactions").select(
            "id, product_id, change_qty, tx_date, tx_type, note, "
            "products(name, units(name)), users(display_name)"
        ).eq("clinic_id", clinic_id).order("tx_date", desc=True)

        if days_filter != "全部":
            days = {"最近 7 天": 7, "最近 30 天": 30, "最近 90 天": 90}[days_filter]
            since = (date.today() - timedelta(days=days)).isoformat()
            query = query.gte("tx_date", since)
        if type_filter != "全部":
            query = query.eq("tx_type", type_filter)

        resp = query.limit(200).execute()

        if not resp.data:
            st.info("沒有異動紀錄")
        else:
            hist_rows = [{
                "日期": t["tx_date"], "品項": t["products"]["name"], "類型": t["tx_type"],
                "數量": float(t["change_qty"]), "單位": t["products"]["units"]["name"],
                "備註": t["note"] or "", "操作人": t["users"]["display_name"] if t.get("users") else "-",
            } for t in resp.data]

            hist_df = pd.DataFrame(hist_rows)
            styled = style_banded(hist_df)
            styled = styled.map(
                lambda v: "color:#DC3545;font-weight:bold" if isinstance(v, (int, float)) and v < 0 else (
                    "color:#28A745" if isinstance(v, (int, float)) and v > 0 else ""),
                subset=["數量"],
            )
            st.dataframe(styled, use_container_width=True, hide_index=True, height=400)

            # 修改/刪除
            st.divider()
            st.subheader("修改或刪除紀錄")

            tx_options = {
                f"#{t['id']} | {t['tx_date']} | {t['products']['name']} | {float(t['change_qty']):+.1f}": t
                for t in resp.data
            }
            selected_tx_key = st.selectbox("選擇紀錄", list(tx_options.keys()), key="tx_hist_select")
            tx_item = tx_options[selected_tx_key]

            col_a, col_b, col_c = st.columns(3)
            with col_a:
                edit_qty = st.number_input("數量", value=float(tx_item["change_qty"]), step=0.1, format="%.1f", key="tx_edit_qty")
            with col_b:
                type_list = ["進貨", "調撥", "廢棄"]
                edit_type = st.selectbox("類型", type_list,
                                         index=type_list.index(tx_item["tx_type"]), key="tx_edit_type")
            with col_c:
                edit_note = st.text_input("備註", value=tx_item["note"] or "", key="tx_edit_note")

            col_save, col_del = st.columns(2)
            with col_save:
                if st.button("💾 修改此筆", type="primary", key="tx_save_btn"):
                    try:
                        sb.table("transactions").update({
                            "change_qty": edit_qty, "tx_type": edit_type,
                            "note": edit_note or None,
                        }).eq("id", tx_item["id"]).execute()
                        recalc_stock(tx_item["product_id"], clinic_id)
                        st.success("已修改")
                        st.rerun()
                    except Exception as e:
                        st.error(f"修改失敗：{e}")

            with col_del:
                if st.button("🗑️ 刪除此筆", type="secondary", key="tx_del_btn"):
                    try:
                        sb.table("transactions").delete().eq("id", tx_item["id"]).execute()
                        recalc_stock(tx_item["product_id"], clinic_id)
                        st.success("已刪除")
                        st.rerun()
                    except Exception as e:
                        st.error(f"刪除失敗：{e}")


# ══════════════════════════════════════════════
#  3. 執行盤點頁面
# ══════════════════════════════════════════════

def page_inventory():
    st.header("📝 執行盤點")

    selected_clinic = st.session_state.get("selected_clinic", "澤豐")
    user = st.session_state.user

    if selected_clinic == "合併檢視":
        st.info("請先選擇單一診所。")
        return

    clinic_id = get_clinic_id(selected_clinic)
    if not clinic_id:
        st.error("找不到該診所")
        return

    sb = get_supabase_client()

    tab_do, tab_photo, tab_print, tab_history = st.tabs(["📝 執行盤點", "📸 照片盤點", "🖨️ 列印盤點表", "📜 盤點歷史"])

    # ── 列印盤點表（含歷史資料）──
    with tab_print:
        st.subheader("🖨️ 列印盤點表")
        st.caption("含歷史盤點/進貨/耗用，最右欄為新盤點數量填寫欄")

        print_cat = st.selectbox("分類", ["全部"] + [c["name"] for c in load_categories()], key="print_cat")

        # 載入品項 + per-clinic 資料
        all_prods = sb.table("products").select(
            "id, name, category_id, categories(name), units(name)"
        ).execute().data
        all_prods = sort_products_by_bopomofo(all_prods)

        cs_print = sb.table("clinic_stock").select(
            "product_id, current_stock, is_active, cabinet, brand1_id, brand2_id"
        ).eq("clinic_id", clinic_id).execute().data
        cs_print_map = {s["product_id"]: s for s in cs_print}

        brands_data = load_brands()
        brand_map_print = {b["id"]: b["name"] for b in brands_data}

        # 載入歷史盤點
        print_logs = sb.table("inventory_logs").select(
            "product_id, current_count_qty, log_date"
        ).eq("clinic_id", clinic_id).order("log_date", desc=True).execute().data

        # 最近 3 個盤點日期
        p_dates = []
        for lg in print_logs:
            if lg["log_date"] not in p_dates:
                p_dates.append(lg["log_date"])
            if len(p_dates) >= 3:
                break
        p_display = list(reversed(p_dates))  # 最舊→最新

        # 每品項盤點
        p_log_map = defaultdict(dict)
        for lg in print_logs:
            pid = lg["product_id"]
            d = lg["log_date"]
            if d not in p_log_map[pid]:
                p_log_map[pid][d] = lg

        # 進退貨
        print_tx = sb.table("transactions").select(
            "product_id, change_qty, tx_date"
        ).eq("clinic_id", clinic_id).execute().data
        tx_by_p = defaultdict(list)
        for tx in print_tx:
            tx_by_p[tx["product_id"]].append(tx)

        active_prods = [p for p in all_prods
                        if cs_print_map.get(p["id"], {}).get("is_active", True)]
        if print_cat != "全部":
            active_prods = [p for p in active_prods if p["categories"]["name"] == print_cat]

        if active_prods:
            # 用 _build_stock_excel 的邏輯產生 Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            thin = Side(style="thin")
            double = Side(style="double")
            border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_font = Font(bold=True, size=10)
            dat_font = Font(size=10)
            ctr = Alignment(horizontal="center", vertical="center")
            lft = Alignment(horizontal="left", vertical="center")

            # 組欄位
            pcols = []  # (header, width, data_key_or_None)
            pcols.append(("注音", 4, "_initial"))
            pcols.append(("品項", 16, "_name"))
            pcols.append(("櫃", 4, "_cabinet"))
            pcols.append(("廠1", 3, "_b1"))
            pcols.append(("廠2", 3, "_b2"))

            for i, d in enumerate(p_display):
                sd = short_date(d)
                pcols.append((sd, 5, f"_inv_{d}"))
                if i < len(p_display) - 1:
                    pcols.append(("進貨", 5, f"_restock_{d}"))
                    pcols.append(("耗用", 5, f"_consume_{d}"))
                else:
                    # 最新盤點後也加進貨+耗用（新生成）
                    pcols.append(("進貨", 5, f"_restock_{d}"))
                    pcols.append(("耗用", 5, f"_consume_{d}"))

            pcols.append(("日期:      ", 10, "_new"))  # 新盤點欄，2倍寬

            buf = io.BytesIO()
            wb = Workbook()
            wb.remove(wb.active)

            MERGE_CATS = {"水藥材", "高貴藥材", "非健保藥材"}
            cat_groups = defaultdict(list)
            for p in active_prods:
                cat_groups[p["categories"]["name"]].append(p)

            def write_print_sheet(ws, prods_list, sheet_title=None):
                ws.page_setup.orientation = "portrait"
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                # 每頁列印都重複第 1 行表頭
                ws.print_title_rows = "1:1"

                # 表頭
                for ci, (h, w, _) in enumerate(pcols, 1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.font = hdr_font
                    cell.alignment = ctr
                    cell.border = border_thin
                    ws.column_dimensions[get_column_letter(ci)].width = w

                row_num = 2
                prev_initial = None
                for p in prods_list:
                    pid = p["id"]
                    cs = cs_print_map.get(pid, {})
                    initial = get_bopomofo_initial(p["name"])

                    if prev_initial is not None and initial != prev_initial:
                        for ci in range(1, len(pcols) + 1):
                            cell = ws.cell(row=row_num - 1, column=ci)
                            cell.border = Border(left=thin, right=thin, top=thin, bottom=double)

                    show_init = initial if initial != prev_initial else ""
                    prev_initial = initial

                    p_logs_data = p_log_map.get(pid, {})

                    for ci, (_, _, key) in enumerate(pcols, 1):
                        if key == "_initial":
                            val = show_init
                        elif key == "_name":
                            val = p["name"]
                        elif key == "_cabinet":
                            val = cs.get("cabinet") or ""
                        elif key == "_b1":
                            val = abbr_brand(brand_map_print.get(cs.get("brand1_id"), "-"))
                        elif key == "_b2":
                            val = abbr_brand(brand_map_print.get(cs.get("brand2_id"), "-"))
                        elif key and key.startswith("_inv_"):
                            d = key[5:]
                            lg = p_logs_data.get(d)
                            val = round(float(lg["current_count_qty"]), 1) if lg else ""
                        elif key and key.startswith("_restock_"):
                            d = key[9:]
                            idx = p_display.index(d) if d in p_display else -1
                            if idx < len(p_display) - 1:
                                next_d = p_display[idx + 1]
                                val = round(sum(float(t["change_qty"]) for t in tx_by_p.get(pid, [])
                                                if d < t["tx_date"] <= next_d), 1)
                            else:
                                # 最新盤點後到今天
                                val = round(sum(float(t["change_qty"]) for t in tx_by_p.get(pid, [])
                                                if t["tx_date"] > d), 1)
                        elif key and key.startswith("_consume_"):
                            d = key[9:]
                            idx = p_display.index(d) if d in p_display else -1
                            lg_this = p_logs_data.get(d)
                            if idx < len(p_display) - 1:
                                next_d = p_display[idx + 1]
                                lg_next = p_logs_data.get(next_d)
                                restock = sum(float(t["change_qty"]) for t in tx_by_p.get(pid, [])
                                              if d < t["tx_date"] <= next_d)
                                if lg_this and lg_next:
                                    val = round(float(lg_this["current_count_qty"]) + restock - float(lg_next["current_count_qty"]), 1)
                                else:
                                    val = ""
                            else:
                                val = ""  # 最新盤點的耗用待新盤點才知道
                        elif key == "_new":
                            val = ""  # 空白待填
                        else:
                            val = ""

                        cell = ws.cell(row=row_num, column=ci, value=val)
                        cell.font = dat_font
                        cell.alignment = lft if key == "_name" else ctr
                        cell.border = border_thin
                    row_num += 1
                return row_num

            # 正常分類各自一頁
            merge_prods = []
            for cat_name, prods in cat_groups.items():
                if cat_name in MERGE_CATS:
                    merge_prods.append((cat_name, prods))
                else:
                    ws = wb.create_sheet(title=str(cat_name)[:31])
                    write_print_sheet(ws, prods)

            # 合併小分類
            if merge_prods:
                ws = wb.create_sheet(title="其他藥材")
                ws.page_setup.orientation = "portrait"
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.print_title_rows = "1:1"
                # 表頭
                for ci, (h, w, _) in enumerate(pcols, 1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.font = hdr_font
                    cell.alignment = ctr
                    cell.border = border_thin
                    ws.column_dimensions[get_column_letter(ci)].width = w
                row_num = 2
                cat_title_f = Font(bold=True, size=11)
                for idx_m, (cat_name, prods) in enumerate(merge_prods):
                    cell = ws.cell(row=row_num, column=1, value=f"【{cat_name}】")
                    cell.font = cat_title_f
                    ws.merge_cells(start_row=row_num, start_column=1,
                                   end_row=row_num, end_column=len(pcols))
                    row_num += 1
                    prev_initial = None
                    for p in prods:
                        pid = p["id"]
                        cs = cs_print_map.get(pid, {})
                        initial = get_bopomofo_initial(p["name"])
                        if prev_initial is not None and initial != prev_initial:
                            for ci in range(1, len(pcols) + 1):
                                cell = ws.cell(row=row_num - 1, column=ci)
                                cell.border = Border(left=thin, right=thin, top=thin, bottom=double)
                        show_init = initial if initial != prev_initial else ""
                        prev_initial = initial
                        p_logs_data = p_log_map.get(pid, {})
                        for ci, (_, _, key) in enumerate(pcols, 1):
                            if key == "_initial": val = show_init
                            elif key == "_name": val = p["name"]
                            elif key == "_cabinet": val = cs.get("cabinet") or ""
                            elif key == "_b1": val = abbr_brand(brand_map_print.get(cs.get("brand1_id"), "-"))
                            elif key == "_b2": val = abbr_brand(brand_map_print.get(cs.get("brand2_id"), "-"))
                            elif key and key.startswith("_inv_"):
                                d = key[5:]
                                lg = p_logs_data.get(d)
                                val = round(float(lg["current_count_qty"]), 1) if lg else ""
                            elif key and key.startswith("_restock_") or key and key.startswith("_consume_") or key == "_new":
                                val = ""
                            else:
                                val = ""
                            cell = ws.cell(row=row_num, column=ci, value=val)
                            cell.font = dat_font
                            cell.alignment = lft if key == "_name" else ctr
                            cell.border = border_thin
                        row_num += 1
                    if idx_m < len(merge_prods) - 1:
                        row_num += 2

            wb.save(buf)
            buf.seek(0)

            st.download_button(
                f"📥 下載盤點表（{len(active_prods)} 品項）",
                data=buf.getvalue(),
                file_name=f"盤點表_{selected_clinic}_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
            )

    # ── 照片盤點 ──
    with tab_photo:
        st.subheader("📸 照片盤點")
        st.caption("上傳手寫盤點表照片，AI 自動辨識品項與數量")

        photo_date = st.date_input("預設盤點日期", value=date.today(), key="photo_date")
        uploaded_files = st.file_uploader(
            "上傳盤點表照片（最多 20 張）",
            type=["jpg", "jpeg", "png", "heic"],
            accept_multiple_files=True,
            key="photo_upload",
        )

        if uploaded_files and len(uploaded_files) > 20:
            st.error("最多上傳 20 張照片")
            uploaded_files = uploaded_files[:20]

        # 初始化 session_state
        if "photo_results" not in st.session_state:
            st.session_state.photo_results = None

        # 步驟 1：辨識
        if uploaded_files and st.button("🔍 開始辨識", type="primary", key="photo_recognize"):
            try:
                from google import genai
                import json as json_mod
                client = genai.Client(api_key=st.secrets["gemini"]["api_key"])

                PROMPT = """【任務目標】
你是一個精準的數據擷取機器人。請從這張手寫盤點表圖檔中提取數據。

【定位與擷取規則】
1. 找到「品項」直行，往下擷取所有藥品名稱（以印刷體為準）。
2. 找到表格最右側標題「日期:」欄位，裡面可能有手寫日期（如 3/25、4/8）。
3. 擷取「日期:」下方每個品項對應的手寫盤點數量。
4. 數字配對規則：
   - 「3+7」等加法請加總（=10）
   - 含小數點的數字照實輸出，如「0.5」→ 0.5、「1.7」→ 1.7
   - 空白、斜線「/」、撇號「-」= null
   - 「K」或「k」= null
   - 有刪除線再重寫的，以最後寫的為準
5. 日期如果看不清或沒寫，date 輸出 null。

【輸出格式】
嚴格 JSON，不要任何解釋文字：
{"date": "4/8", "items": [{"name": "加味逍遙散", "qty": 15}, {"name": "附子", "qty": 0.5}, {"name": "葛根湯", "qty": null}]}"""

                all_items = []
                detected_dates = []
                progress = st.progress(0, text="辨識中...")

                for i, f in enumerate(uploaded_files):
                    progress.progress((i + 1) / len(uploaded_files),
                                      text=f"辨識第 {i+1}/{len(uploaded_files)} 張...")
                    img_bytes = f.read()
                    response = client.models.generate_content(
                        model="gemini-2.5-flash",
                        contents=[
                            genai.types.Part.from_bytes(
                                data=img_bytes, mime_type=f.type or "image/jpeg"),
                            PROMPT,
                        ],
                    )
                    text = response.text.strip()
                    if text.startswith("```"):
                        text = text.split("\n", 1)[1].rsplit("```", 1)[0].strip()
                    result = json_mod.loads(text)
                    if result.get("date"):
                        detected_dates.append(result["date"])
                    for item in result.get("items", []):
                        if item.get("name"):
                            all_items.append(item)

                progress.empty()

                # 日期解析
                unique_dates = list(set(d for d in detected_dates if d))
                parsed_ai_dates = set()
                for d in unique_dates:
                    try:
                        parts = d.replace(".", "/").replace("-", "/").split("/")
                        m, dy = int(parts[0]), int(parts[1])
                        y = date.today().year
                        parsed_ai_dates.add(date(y, m, dy).isoformat())
                    except Exception:
                        pass

                ai_date = list(parsed_ai_dates)[0] if len(parsed_ai_dates) == 1 else None

                # 品名匹配
                all_products_db = sb.table("products").select("id, name").execute().data
                name_to_pid = {}
                for p in all_products_db:
                    name_to_pid[p["name"]] = p["id"]
                    name_to_pid[p["name"].strip()] = p["id"]
                    name_to_pid[p["name"].replace(" ", "")] = p["id"]

                cs_photo = sb.table("clinic_stock").select(
                    "product_id, current_stock, is_active"
                ).eq("clinic_id", clinic_id).execute().data
                active_pids = {s["product_id"] for s in cs_photo if s.get("is_active", True)}
                stock_map_photo = {s["product_id"]: float(s["current_stock"]) for s in cs_photo}

                # 去重
                item_map = {}
                for item in all_items:
                    name = item["name"].strip()
                    item_map[name] = item.get("qty")

                matched, unmatched = [], []
                for name, qty in item_map.items():
                    pid = name_to_pid.get(name) or name_to_pid.get(name.replace(" ", ""))
                    if pid and pid in active_pids:
                        matched.append({
                            "product_id": int(pid),
                            "品項": next(p["name"] for p in all_products_db if p["id"] == pid),
                            "辨識數量": round(float(qty), 1) if qty is not None else None,
                            "帳面庫存": stock_map_photo.get(pid, 0),
                        })
                    else:
                        unmatched.append({"品項(未匹配)": name, "辨識數量": qty})

                # 存入 session_state
                st.session_state.photo_results = {
                    "matched": matched,
                    "unmatched": unmatched,
                    "ai_date": ai_date,
                    "parsed_ai_dates": sorted(parsed_ai_dates) if parsed_ai_dates else [],
                    "photo_count": len(uploaded_files),
                    "item_count": len(all_items),
                }
                st.rerun()

            except Exception as e:
                st.error(f"辨識失敗：{e}")
                import traceback
                st.code(traceback.format_exc())

        # 步驟 2：顯示辨識結果（從 session_state 讀取）
        results = st.session_state.get("photo_results")
        if results:
            matched = results["matched"]
            unmatched = results["unmatched"]
            ai_date = results["ai_date"]

            st.success(f"辨識完成！{results['photo_count']} 張照片，{results['item_count']} 筆資料")

            # 日期處理
            if ai_date and ai_date != str(photo_date):
                st.warning(f"照片日期 **{ai_date}** 與選擇日期 **{photo_date}** 不同")
                date_choice = st.radio("使用哪個日期？",
                    [f"照片日期 ({ai_date})", f"手選日期 ({photo_date})"],
                    key="photo_date_choice")
                final_date = ai_date if "照片" in date_choice else str(photo_date)
            elif ai_date:
                final_date = ai_date
                st.info(f"照片辨識日期：{ai_date}")
            elif results["parsed_ai_dates"] and len(results["parsed_ai_dates"]) > 1:
                st.warning(f"多張照片辨識到不同日期")
                final_date = st.selectbox("請選擇正確日期",
                    [str(photo_date)] + results["parsed_ai_dates"], key="photo_multi_date")
            else:
                final_date = str(photo_date)
                st.info(f"使用選擇日期：{photo_date}")

            st.markdown(f"**匹配成功 {len(matched)} 筆** / 未匹配 {len(unmatched)} 筆")

            if unmatched:
                with st.expander(f"⚠️ {len(unmatched)} 筆未匹配"):
                    st.dataframe(pd.DataFrame(unmatched), use_container_width=True, hide_index=True)

            if matched:
                match_df = pd.DataFrame(matched)

                st.subheader(f"📋 預覽辨識結果（日期：{final_date}）")
                edited_match = st.data_editor(
                    match_df[["品項", "帳面庫存", "辨識數量"]],
                    use_container_width=True, hide_index=True,
                    column_config={
                        "品項": st.column_config.TextColumn(disabled=True),
                        "帳面庫存": st.column_config.NumberColumn(disabled=True, format="%.1f"),
                        "辨識數量": st.column_config.NumberColumn("盤點數量 ✏️", min_value=0, format="%.1f"),
                    },
                    height=min(len(match_df) * 35 + 38, 500),
                    key="photo_result_editor",
                )

                filled_count = edited_match["辨識數量"].notna().sum()
                st.caption(f"有效數量 {filled_count} / {len(edited_match)}")

                col_save, col_clear = st.columns([3, 1])
                with col_save:
                    if st.button("✅ 確認存檔盤點", type="primary", key="photo_save",
                                 disabled=(filled_count == 0)):
                        entries = []
                        for idx, row in edited_match.iterrows():
                            if pd.notna(row["辨識數量"]):
                                entries.append((match_df.iloc[idx]["product_id"],
                                               round(float(row["辨識數量"]), 1)))

                        if not entries:
                            st.error("沒有有效的盤點數量")
                        else:
                            dup_check = sb.table("inventory_sessions").select("id").eq(
                                "clinic_id", clinic_id).eq("session_date", final_date).execute().data
                            has_dup = False
                            for ds in dup_check:
                                lc = sb.table("inventory_logs").select("id").eq(
                                    "session_id", ds["id"]).limit(1).execute().data
                                if lc:
                                    has_dup = True
                                    break

                            if has_dup:
                                st.warning(f"⚠️ {final_date} 已有盤點紀錄，請至「盤點歷史」修改。")
                            else:
                                try:
                                    # 重新查詢（避免 rerun 後變數消失）
                                    cs_save = sb.table("clinic_stock").select(
                                        "product_id, current_stock"
                                    ).eq("clinic_id", clinic_id).execute().data
                                    stock_map_save = {s["product_id"]: float(s["current_stock"]) for s in cs_save}

                                    last_logs = sb.table("inventory_logs").select(
                                        "product_id, current_count_qty, log_date"
                                    ).eq("clinic_id", clinic_id).order(
                                        "log_date", desc=True).execute().data
                                    last_count_map, last_date_map = {}, {}
                                    for lg in last_logs:
                                        pid = lg["product_id"]
                                        if pid not in last_count_map:
                                            last_count_map[pid] = float(lg["current_count_qty"])
                                            last_date_map[pid] = lg["log_date"]

                                    all_tx = sb.table("transactions").select(
                                        "product_id, change_qty, tx_date"
                                    ).eq("clinic_id", clinic_id).execute().data
                                    tx_by_pid = defaultdict(list)
                                    for tx in all_tx:
                                        tx_by_pid[tx["product_id"]].append(tx)

                                    session_resp = sb.table("inventory_sessions").insert({
                                        "clinic_id": int(clinic_id),
                                        "session_date": final_date,
                                        "operator_id": int(user["id"]),
                                        "status": "已完成",
                                        "completed_at": datetime.now().isoformat(),
                                    }).execute()
                                    session_id = session_resp.data[0]["id"]

                                    logs_to_insert = []
                                    for product_id, count_qty in entries:
                                        last_qty = last_count_map.get(product_id,
                                            stock_map_save.get(product_id, 0))
                                        last_dt = last_date_map.get(product_id, "1900-01-01")
                                        restock_sum = round(sum(
                                            float(t["change_qty"]) for t in tx_by_pid.get(product_id, [])
                                            if last_dt < t["tx_date"] <= final_date
                                        ), 1)
                                        consumed = round(last_qty + restock_sum - count_qty, 1)
                                        logs_to_insert.append({
                                            "session_id": int(session_id),
                                            "product_id": int(product_id),
                                            "clinic_id": int(clinic_id),
                                            "last_count_qty": last_qty,
                                            "restock_qty_since_last": restock_sum,
                                            "current_count_qty": count_qty,
                                            "consumed_qty": consumed,
                                            "log_date": final_date,
                                        })

                                    sb.table("inventory_logs").insert(logs_to_insert).execute()
                                    for entry in logs_to_insert:
                                        recalc_stock(entry["product_id"], int(clinic_id))

                                    st.session_state.photo_results = None
                                    st.success(f"照片盤點完成！{len(logs_to_insert)} 個品項（日期：{final_date}）")
                                    st.balloons()

                                except Exception as e:
                                    st.error(f"存檔失敗：{e}")

                with col_clear:
                    if st.button("🗑️ 清除結果", key="photo_clear"):
                        st.session_state.photo_results = None
                        st.rerun()

    # ── 執行盤點 ──
    with tab_do:
        inv_date = st.date_input("盤點日期（可選過去日期回填）", value=date.today(), key="inv_date")

        # 檢查重複日期
        existing_sessions = sb.table("inventory_sessions").select("id").eq(
            "clinic_id", clinic_id).eq("session_date", str(inv_date)).execute().data
        # 只檢查有 logs 的 sessions
        has_logs = False
        for es in existing_sessions:
            log_check = sb.table("inventory_logs").select("id").eq("session_id", es["id"]).limit(1).execute().data
            if log_check:
                has_logs = True
                break

        if has_logs:
            st.warning(f"⚠️ {inv_date} 已有盤點紀錄，請至「盤點歷史」修改。")
        else:
            categories = load_categories()
            cat_filter = st.selectbox("分類（建議逐分類盤點）", ["全部"] + [c["name"] for c in categories], key="inv_cat")

            products = sb.table("products").select(
                "id, name, category_id, categories(name), units(name)"
            ).execute().data
            products = sort_products_by_bopomofo(products)

            # per-clinic 啟用篩選
            cs_data = sb.table("clinic_stock").select(
                "product_id, current_stock, is_active"
            ).eq("clinic_id", clinic_id).execute().data
            cs_map = {s["product_id"]: s for s in cs_data}
            products = [p for p in products if cs_map.get(p["id"], {}).get("is_active", True)]

            stock_map = {s["product_id"]: float(s["current_stock"]) for s in cs_data}

            filtered = products
            if cat_filter != "全部":
                filtered = [p for p in filtered if p["categories"]["name"] == cat_filter]

            if not filtered:
                st.info("沒有符合條件的品項")
            else:
                edit_data = [{
                    "product_id": p["id"], "品項名稱": p["name"], "分類": p["categories"]["name"],
                    "單位": p["units"]["name"], "帳面庫存": stock_map.get(p["id"], 0), "盤點數量": None,
                } for p in filtered]

                df = pd.DataFrame(edit_data)
                st.markdown(f"**{selected_clinic}** — 盤點日期：**{inv_date}** — 共 {len(df)} 個品項")

                edited_df = st.data_editor(
                    df[["品項名稱", "分類", "單位", "帳面庫存", "盤點數量"]],
                    use_container_width=True, hide_index=True, num_rows="fixed",
                    column_config={
                        "品項名稱": st.column_config.TextColumn(disabled=True),
                        "分類": st.column_config.TextColumn(disabled=True),
                        "單位": st.column_config.TextColumn(disabled=True),
                        "帳面庫存": st.column_config.NumberColumn(disabled=True, format="%.1f"),
                        "盤點數量": st.column_config.NumberColumn("盤點數量 ✏️", min_value=0, format="%.1f"),
                    },
                    height=min(len(df) * 35 + 38, 600),
                    key="inv_editor",
                )

                filled = edited_df["盤點數量"].notna().sum()
                st.caption(f"已填 {filled} / {len(edited_df)}")

                if st.button("✅ 確認盤點", type="primary", disabled=(filled == 0)):
                    # 先收集所有有填數量的項目，避免建立空批次
                    entries = []
                    for idx, row in edited_df.iterrows():
                        if pd.notna(row["盤點數量"]):
                            entries.append((idx, round(float(row["盤點數量"]), 1)))

                    if not entries:
                        st.error("沒有填入任何盤點數量，請重新填寫後再送出。")
                    else:
                        try:
                            # 取得每個品項的最新盤點
                            last_logs = sb.table("inventory_logs").select(
                                "product_id, current_count_qty, log_date"
                            ).eq("clinic_id", clinic_id).order("log_date", desc=True).execute().data

                            last_count_map, last_date_map = {}, {}
                            for log in last_logs:
                                pid = log["product_id"]
                                if pid not in last_count_map:
                                    last_count_map[pid] = float(log["current_count_qty"])
                                    last_date_map[pid] = log["log_date"]

                            all_tx = sb.table("transactions").select(
                                "product_id, change_qty, tx_date"
                            ).eq("clinic_id", clinic_id).execute().data

                            tx_by_pid = defaultdict(list)
                            for tx in all_tx:
                                tx_by_pid[tx["product_id"]].append(tx)

                            # 資料都準備好了，才建立 session
                            session_resp = sb.table("inventory_sessions").insert({
                                "clinic_id": clinic_id, "session_date": str(inv_date),
                                "operator_id": user["id"], "status": "已完成",
                                "completed_at": datetime.now().isoformat(),
                            }).execute()
                            session_id = session_resp.data[0]["id"]

                            logs_to_insert, results = [], []

                            for idx, count_qty in entries:
                                product_id = int(df.iloc[idx]["product_id"])
                                last_qty = round(float(last_count_map.get(product_id, stock_map.get(product_id, 0))), 1)
                                last_dt = last_date_map.get(product_id, "1900-01-01")

                                restock_sum = round(sum(
                                    float(t["change_qty"]) for t in tx_by_pid.get(product_id, [])
                                    if last_dt < t["tx_date"] <= str(inv_date)
                                ), 1)
                                consumed = round(last_qty + restock_sum - count_qty, 1)

                                logs_to_insert.append({
                                    "session_id": int(session_id), "product_id": product_id,
                                    "clinic_id": int(clinic_id),
                                    "last_count_qty": last_qty,
                                    "restock_qty_since_last": restock_sum,
                                    "current_count_qty": count_qty,
                                    "consumed_qty": consumed,
                                    "log_date": str(inv_date),
                                })

                                results.append({
                                    "品項": edited_df.iloc[idx]["品項名稱"],
                                    "上次盤點": last_qty,
                                    "期間進貨": restock_sum, "本次盤點": count_qty, "耗用量": consumed,
                                })

                            # 先寫入 logs，再重算庫存（recalc 要抓到新寫入的盤點）
                            sb.table("inventory_logs").insert(logs_to_insert).execute()

                            for entry in logs_to_insert:
                                recalc_stock(entry["product_id"], int(clinic_id))

                            st.success(f"盤點完成！共 {len(results)} 個品項（日期：{inv_date}）")
                            if results:
                                st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)

                        except Exception as e:
                            st.error(f"盤點失敗：{e}")

    # ── 盤點歷史 ──
    with tab_history:
        sessions = sb.table("inventory_sessions").select(
            "id, session_date, status, users(display_name)"
        ).eq("clinic_id", clinic_id).order("session_date", desc=True).limit(20).execute().data

        if not sessions:
            st.info("尚無盤點紀錄")
        else:
            # 載入全品項（用於顯示完整清單）
            all_products = sb.table("products").select(
                "id, name, category_id, units(name)"
            ).execute().data
            all_products = sort_products_by_bopomofo(all_products)

            # per-clinic 啟用
            cs_active = sb.table("clinic_stock").select(
                "product_id, current_stock, is_active"
            ).eq("clinic_id", clinic_id).execute().data
            active_pids = {s["product_id"] for s in cs_active if s.get("is_active", True)}
            stock_map_hist = {s["product_id"]: float(s["current_stock"]) for s in cs_active}
            all_products = [p for p in all_products if p["id"] in active_pids]

            for s in sessions:
                operator = s["users"]["display_name"] if s.get("users") else "-"
                with st.expander(f"📋 {s['session_date']}（{operator}）", expanded=False):
                    # 載入此 session 已有的 logs
                    logs = sb.table("inventory_logs").select(
                        "id, product_id, last_count_qty, restock_qty_since_last, "
                        "current_count_qty, consumed_qty"
                    ).eq("session_id", s["id"]).execute().data
                    log_by_pid = {l["product_id"]: l for l in logs}

                    counted = len(logs)
                    total = len(all_products)
                    st.caption(f"已盤 {counted} / {total} 個品項")

                    # 組裝全品項表格（已盤的顯示數量，未盤的留空）
                    log_rows = []
                    row_meta = []  # (product_id, log_id_or_None)
                    for p in all_products:
                        pid = p["id"]
                        log = log_by_pid.get(pid)
                        if log:
                            log_rows.append({
                                "品項": p["name"],
                                "分類": "",
                                "單位": p["units"]["name"],
                                "帳面庫存": stock_map_hist.get(pid, 0),
                                "盤點數量": float(log["current_count_qty"]) if log["current_count_qty"] is not None else None,
                            })
                            row_meta.append((pid, log["id"]))
                        else:
                            log_rows.append({
                                "品項": p["name"],
                                "分類": "",
                                "單位": p["units"]["name"],
                                "帳面庫存": stock_map_hist.get(pid, 0),
                                "盤點數量": None,
                            })
                            row_meta.append((pid, None))

                    log_df = pd.DataFrame(log_rows)

                    edited_log_df = st.data_editor(
                        log_df,
                        use_container_width=True, hide_index=True,
                        column_config={
                            "品項": st.column_config.TextColumn(disabled=True),
                            "分類": st.column_config.TextColumn(disabled=True),
                            "單位": st.column_config.TextColumn(disabled=True),
                            "帳面庫存": st.column_config.NumberColumn(disabled=True, format="%.1f"),
                            "盤點數量": st.column_config.NumberColumn("盤點數量 ✏️", min_value=0, format="%.1f"),
                        },
                        height=min(len(log_df) * 35 + 38, 600),
                        key=f"hist_editor_{s['id']}",
                    )

                    col_save, col_del = st.columns([3, 1])
                    with col_save:
                        if st.button("💾 儲存修改", key=f"hist_save_{s['id']}", type="primary"):
                            # 載入進貨資料
                            all_tx = sb.table("transactions").select(
                                "product_id, change_qty, tx_date"
                            ).eq("clinic_id", clinic_id).execute().data
                            tx_by_pid = defaultdict(list)
                            for tx in all_tx:
                                tx_by_pid[tx["product_id"]].append(tx)

                            # 載入每品項最新盤點（排除當前 session）
                            prev_logs = sb.table("inventory_logs").select(
                                "product_id, current_count_qty, log_date"
                            ).eq("clinic_id", clinic_id).neq(
                                "session_id", s["id"]
                            ).order("log_date", desc=True).execute().data
                            prev_map = {}
                            prev_date_map = {}
                            for pl in prev_logs:
                                pid = pl["product_id"]
                                if pid not in prev_map:
                                    prev_map[pid] = float(pl["current_count_qty"])
                                    prev_date_map[pid] = pl["log_date"]

                            updated, inserted = 0, 0
                            for idx in range(len(edited_log_df)):
                                new_val = edited_log_df.iloc[idx]["盤點數量"]
                                old_val = log_df.iloc[idx]["盤點數量"]
                                pid, log_id = row_meta[idx]

                                if pd.isna(new_val):
                                    continue

                                new_qty = round(float(new_val), 1)

                                # 計算耗用
                                last_qty = round(float(prev_map.get(pid, stock_map_hist.get(pid, 0))), 1)
                                last_dt = prev_date_map.get(pid, "1900-01-01")
                                restock_sum = round(sum(
                                    float(t["change_qty"]) for t in tx_by_pid.get(pid, [])
                                    if last_dt < t["tx_date"] <= s["session_date"]
                                ), 1)
                                consumed = round(last_qty + restock_sum - new_qty, 1)

                                if log_id is not None:
                                    # 已有 log — 更新
                                    if pd.notna(old_val) and round(float(old_val), 1) == new_qty:
                                        continue
                                    sb.table("inventory_logs").update({
                                        "current_count_qty": new_qty,
                                        "consumed_qty": consumed,
                                    }).eq("id", log_id).execute()
                                    updated += 1
                                else:
                                    # 新增 log
                                    sb.table("inventory_logs").insert({
                                        "session_id": int(s["id"]),
                                        "product_id": int(pid),
                                        "clinic_id": int(clinic_id),
                                        "last_count_qty": last_qty,
                                        "restock_qty_since_last": restock_sum,
                                        "current_count_qty": new_qty,
                                        "consumed_qty": consumed,
                                        "log_date": s["session_date"],
                                    }).execute()
                                    inserted += 1

                                # 重算即時庫存
                                recalc_stock(int(pid), int(clinic_id))

                            msg = []
                            if updated > 0:
                                msg.append(f"修改 {updated} 筆")
                            if inserted > 0:
                                msg.append(f"新增 {inserted} 筆")
                            if msg:
                                st.success("已" + "、".join(msg))
                                st.rerun()
                            else:
                                st.info("沒有變更")

                    with col_del:
                        if user["role"] == "admin":
                            if st.button("🗑️ 刪除整筆", key=f"hist_del_{s['id']}", type="secondary"):
                                try:
                                    # 收集受影響的品項
                                    affected_pids = set(l["product_id"] for l in logs)

                                    # 先刪除 logs 和 session
                                    sb.table("inventory_logs").delete().eq("session_id", s["id"]).execute()
                                    sb.table("inventory_sessions").delete().eq("id", s["id"]).execute()

                                    # 再重算每個受影響品項的庫存
                                    for pid in affected_pids:
                                        recalc_stock(int(pid), int(clinic_id))

                                    st.success("已刪除整筆盤點")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"刪除失敗：{e}")


# ══════════════════════════════════════════════
#  4. 品項管理（per-clinic 廠牌/啟用）
# ══════════════════════════════════════════════

def page_items():
    st.header("📋 品項管理")

    selected_clinic = st.session_state.get("selected_clinic", "澤豐")

    if selected_clinic == "合併檢視":
        st.info("請先選擇單一診所來管理品項。")
        return

    clinic_id = get_clinic_id(selected_clinic)
    if not clinic_id:
        st.error("找不到該診所")
        return

    products = load_products()
    brands = load_brands()
    categories = load_categories()
    units = load_units()

    brand_map = {b["id"]: b["name"] for b in brands}
    unit_map = {u["id"]: u["name"] for u in units}

    sb = get_supabase_client()

    # 載入 per-clinic 資料
    cs_data = sb.table("clinic_stock").select(
        "product_id, brand1_id, brand2_id, is_active, cabinet"
    ).eq("clinic_id", clinic_id).execute().data
    cs_map = {s["product_id"]: s for s in cs_data}

    tab_list, tab_add = st.tabs(["📊 品項清單", "➕ 新增品項"])

    with tab_list:
        st.caption(f"📍 顯示 **{selected_clinic}** 的廠牌與啟用狀態")

        col1, col2, col3 = st.columns([2, 2, 3])
        with col1:
            cat_filter = st.selectbox("分類", ["全部"] + [c["name"] for c in categories], key="item_cat")
        with col2:
            active_filter = st.selectbox("狀態", ["啟用中", "已停用", "全部"], key="item_active")
        with col3:
            search = st.text_input("搜尋", placeholder="中文 或 注音首碼", key="item_search")

        all_names = tuple(p["name"] for p in products)
        key_index = build_bopomofo_index(all_names)

        rows, product_ids = [], []
        cat_names = [c["name"] for c in categories]
        unit_names = [u["name"] for u in units]
        brand_names = ["-"] + [b["name"] for b in brands]

        for p in products:
            cs = cs_map.get(p["id"], {})
            is_active = cs.get("is_active", True)

            if cat_filter != "全部" and p["categories"]["name"] != cat_filter:
                continue
            if active_filter == "啟用中" and not is_active:
                continue
            if active_filter == "已停用" and is_active:
                continue
            if search and not match_search(p["name"], key_index.get(p["name"], ("", "")), search):
                continue

            rows.append({
                "品項名稱": p["name"], "分類": p["categories"]["name"], "單位": p["units"]["name"],
                "櫃位": cs.get("cabinet") or "",
                "第一廠牌": brand_map.get(cs.get("brand1_id"), "-"),
                "第二廠牌": brand_map.get(cs.get("brand2_id"), "-"),
                "規格": p["spec_note"] or "-", "啟用": is_active,
            })
            product_ids.append(p["id"])

        if rows:
            df = pd.DataFrame(rows)
            can_edit = st.session_state.user["role"] in ("admin", "manager")

            if can_edit:
                edited_df = st.data_editor(
                    df, use_container_width=True, hide_index=True,
                    height=min(len(df) * 35 + 38, 600),
                    column_config={
                        "品項名稱": st.column_config.TextColumn(),
                        "分類": st.column_config.SelectboxColumn(options=cat_names),
                        "單位": st.column_config.SelectboxColumn(options=unit_names),
                        "櫃位": st.column_config.TextColumn(),
                        "第一廠牌": st.column_config.SelectboxColumn(options=brand_names),
                        "第二廠牌": st.column_config.SelectboxColumn(options=brand_names),
                        "規格": st.column_config.TextColumn(),
                        "啟用": st.column_config.CheckboxColumn(),
                    },
                    key="items_editor",
                )

                if st.button("💾 儲存修改", type="primary"):
                    updated = 0
                    for idx in range(len(edited_df)):
                        old, new = df.iloc[idx], edited_df.iloc[idx]
                        pid = product_ids[idx]

                        # 共用欄位 → 更新 products 表
                        product_changes = {}
                        if old["品項名稱"] != new["品項名稱"]:
                            product_changes["name"] = new["品項名稱"]
                        if old["分類"] != new["分類"]:
                            product_changes["category_id"] = next(c["id"] for c in categories if c["name"] == new["分類"])
                        if old["單位"] != new["單位"]:
                            product_changes["unit_id"] = next(u["id"] for u in units if u["name"] == new["單位"])
                        if old["規格"] != new["規格"]:
                            product_changes["spec_note"] = new["規格"] if new["規格"] != "-" else None

                        if product_changes:
                            try:
                                sb.table("products").update(product_changes).eq("id", pid).execute()
                                updated += 1
                            except Exception as e:
                                err_msg = str(e)
                                if "duplicate" in err_msg.lower() or "unique" in err_msg.lower():
                                    st.error(f"品項「{new['品項名稱']}」名稱重複，無法更新")
                                else:
                                    st.error(f"更新「{old['品項名稱']}」失敗：{e}")
                                continue

                        # per-clinic 欄位 → 更新 clinic_stock 表
                        cs_changes = {}
                        if old.get("櫃位", "") != new.get("櫃位", ""):
                            cs_changes["cabinet"] = new["櫃位"] if new["櫃位"] else None
                        if old["第一廠牌"] != new["第一廠牌"]:
                            cs_changes["brand1_id"] = next((b["id"] for b in brands if b["name"] == new["第一廠牌"]), None)
                        if old["第二廠牌"] != new["第二廠牌"]:
                            cs_changes["brand2_id"] = next((b["id"] for b in brands if b["name"] == new["第二廠牌"]), None)
                        if old["啟用"] != new["啟用"]:
                            cs_changes["is_active"] = bool(new["啟用"])

                        if cs_changes:
                            sb.table("clinic_stock").update(cs_changes).eq(
                                "product_id", pid).eq("clinic_id", clinic_id).execute()
                            updated += 1

                    if updated > 0:
                        st.success(f"已更新 {updated} 個品項")
                        load_products.clear()
                    else:
                        st.info("沒有變更")

                st.caption(f"共 {len(df)} 個品項（停用品項=軟刪除，取消勾選「啟用」即可）— {selected_clinic}")
            else:
                st.dataframe(style_banded(df), use_container_width=True, hide_index=True, height=min(len(df) * 35 + 38, 600))
                st.caption(f"共 {len(df)} 個品項 — {selected_clinic}")

    with tab_add:
        if st.session_state.user["role"] not in ("admin", "manager"):
            st.warning("僅管理者可新增品項")
            return

        with st.form("add_product_form"):
            st.subheader("新增品項")
            col1, col2 = st.columns(2)
            with col1:
                new_name = st.text_input("品項名稱 *")
                new_cat = st.selectbox("分類 *", [c["name"] for c in categories])
                selected_cat = next((c for c in categories if c["name"] == new_cat), None)
            with col2:
                unit_options = [u["name"] for u in units]
                default_idx = 0
                if selected_cat and selected_cat["default_unit_id"]:
                    dn = unit_map.get(selected_cat["default_unit_id"], "罐")
                    if dn in unit_options:
                        default_idx = unit_options.index(dn)
                new_unit = st.selectbox("單位", unit_options, index=default_idx)
                new_spec = st.text_input("規格", value=selected_cat["default_spec_note"] or "" if selected_cat else "")

            col3, col4 = st.columns(2)
            with col3:
                bo = ["-"] + [b["name"] for b in brands]
                new_b1 = st.selectbox("第一廠牌（叫貨首選）", bo)
            with col4:
                new_b2 = st.selectbox("第二廠牌（缺藥備選，可留空）", bo, key="add_b2")

            if st.form_submit_button("新增品項", type="primary", use_container_width=True):
                if not new_name.strip():
                    st.error("請輸入品項名稱")
                else:
                    try:
                        b1_id = next((b["id"] for b in brands if b["name"] == new_b1), None)
                        b2_id = next((b["id"] for b in brands if b["name"] == new_b2), None)

                        resp = sb.table("products").insert({
                            "name": new_name.strip(),
                            "category_id": next(c["id"] for c in categories if c["name"] == new_cat),
                            "unit_id": next(u["id"] for u in units if u["name"] == new_unit),
                            "brand1_id": b1_id, "brand2_id": b2_id,
                            "spec_note": new_spec or None, "is_active": True,
                        }).execute()
                        pid = resp.data[0]["id"]

                        # 為兩間診所建立 clinic_stock（含廠牌）
                        for c in sb.table("clinics").select("id").execute().data:
                            sb.table("clinic_stock").insert({
                                "product_id": pid, "clinic_id": c["id"],
                                "current_stock": 0,
                                "brand1_id": b1_id, "brand2_id": b2_id,
                                "is_active": True,
                            }).execute()

                        st.success(f"已新增：{new_name}")
                        load_products.clear()
                    except Exception as e:
                        if "duplicate" in str(e).lower():
                            st.error(f"品項「{new_name}」已存在")
                        else:
                            st.error(f"新增失敗：{e}")


# ══════════════════════════════════════════════
#  5. 數據分析
# ══════════════════════════════════════════════

def page_analytics():
    st.header("📊 數據分析")

    selected_clinic = st.session_state.get("selected_clinic", "澤豐")
    sb = get_supabase_client()

    if selected_clinic == "合併檢視":
        clinic_ids = [1, 2]
    else:
        cid = get_clinic_id(selected_clinic)
        clinic_ids = [cid] if cid else []

    all_logs = []
    for cid in clinic_ids:
        resp = sb.table("inventory_logs").select(
            "product_id, consumed_qty, products(name, category_id, categories(name), units(name))"
        ).eq("clinic_id", cid).order("log_date", desc=True).execute()
        all_logs.extend(resp.data)

    if not all_logs:
        st.info("尚無盤點資料。")
        return

    product_consumed = defaultdict(list)
    product_info = {}
    for log in all_logs:
        pid = log["product_id"]
        if log["consumed_qty"] is not None and log["consumed_qty"] > 0:
            product_consumed[pid].append(float(log["consumed_qty"]))
        product_info[pid] = {
            "name": log["products"]["name"],
            "category": log["products"]["categories"]["name"],
            "unit": log["products"]["units"]["name"],
        }

    tab_reorder, tab_ranking, tab_cabinet = st.tabs(["📦 建議叫貨", "🏆 常用排名", "🗄️ 櫃位分類"])

    with tab_reorder:
        settings = sb.table("system_settings").select("*").execute().data
        safety = float(next(s["value"] for s in settings if s["key"] == "safety_factor"))
        multiplier = float(next(s["value"] for s in settings if s["key"] == "stock_target_multiplier"))

        stock_data = []
        for cid in clinic_ids:
            resp = sb.table("clinic_stock").select("product_id, current_stock, clinic_id").eq("clinic_id", cid).execute()
            stock_data.extend(resp.data)

        stock_map = {(s["product_id"], s["clinic_id"]): float(s["current_stock"]) for s in stock_data}

        reorder_rows = []
        for pid, consumed_list in product_consumed.items():
            avg = sum(consumed_list) / len(consumed_list)
            if avg <= 0:
                continue
            info = product_info[pid]
            for cid in clinic_ids:
                current = stock_map.get((pid, cid), 0)
                if current < avg * safety:
                    reorder_rows.append({
                        "診所": "澤豐" if cid == 1 else "澤沛",
                        "品項": info["name"], "分類": info["category"],
                        "目前庫存": current, "平均耗用": round(avg, 1),
                        "建議叫貨": max(0, round(avg * multiplier - current, 1)),
                    })

        if reorder_rows:
            reorder_df = pd.DataFrame(reorder_rows).sort_values(["診所", "分類"]).reset_index(drop=True)
            st.dataframe(style_banded(reorder_df),
                         use_container_width=True, hide_index=True)
        else:
            st.success("所有品項庫存充足")

    with tab_ranking:
        categories = load_categories()
        rank_cat = st.selectbox("分類", ["全部"] + [c["name"] for c in categories], key="rank_cat")

        ranking = []
        for pid, consumed_list in product_consumed.items():
            info = product_info[pid]
            if rank_cat != "全部" and info["category"] != rank_cat:
                continue
            total = sum(consumed_list)
            if total > 0:
                ranking.append({"品項": info["name"], "分類": info["category"],
                                "總耗用量": round(total, 1), "平均耗用": round(total / len(consumed_list), 1)})

        if ranking:
            rank_df = pd.DataFrame(ranking).sort_values("總耗用量", ascending=False).reset_index(drop=True)
            st.bar_chart(rank_df.head(15).set_index("品項")[["總耗用量"]], color="#6A5ACD")
            st.dataframe(style_banded(rank_df), use_container_width=True, hide_index=True)

    with tab_cabinet:
        if selected_clinic == "合併檢視":
            st.info("請先選擇單一診所。")
        else:
            cab_clinic_id = get_clinic_id(selected_clinic)
            cab_sb = get_supabase_client()
            cab_products = cab_sb.table("products").select(
                "id, name, category_id, categories(name), units(name)"
            ).execute().data

            cab_cs = cab_sb.table("clinic_stock").select(
                "product_id, current_stock, cabinet, is_active, brand1_id, brand2_id"
            ).eq("clinic_id", cab_clinic_id).execute().data
            cab_cs_map = {s["product_id"]: s for s in cab_cs}

            cab_brands = load_brands()
            cab_brand_map = {b["id"]: b["name"] for b in cab_brands}

            cab_rows = []
            for p in cab_products:
                cs = cab_cs_map.get(p["id"])
                if not cs or not cs.get("is_active", True):
                    continue
                cabinet_val = cs.get("cabinet") or ""
                cab_rows.append({
                    "櫃位": cabinet_val,
                    "品項": p["name"],
                    "分類": p["categories"]["name"],
                    "單位": p["units"]["name"],
                    "廠1": abbr_brand(cab_brand_map.get(cs.get("brand1_id"), "-")),
                    "廠2": abbr_brand(cab_brand_map.get(cs.get("brand2_id"), "-")),
                    "庫存": float(cs["current_stock"]),
                })

            if cab_rows:
                cab_df = pd.DataFrame(cab_rows)

                # 選分類篩選
                cab_cats = sorted(cab_df["分類"].unique())
                cab_cat_filter = st.selectbox("分類", ["全部"] + list(cab_cats), key="cab_cat")
                if cab_cat_filter != "全部":
                    cab_df = cab_df[cab_df["分類"] == cab_cat_filter]

                # 依櫃位排序
                cab_df = cab_df.sort_values(["櫃位", "品項"]).reset_index(drop=True)
                st.dataframe(style_banded(cab_df), use_container_width=True, hide_index=True)

                # 匯出
                if st.button("📥 匯出櫃位分類表", type="primary"):
                    from openpyxl.styles import Font, Border, Side, Alignment
                    from openpyxl import Workbook
                    from openpyxl.utils import get_column_letter

                    thin_s = Side(style="thin")
                    double_s = Side(style="double")
                    border_t = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

                    wb = Workbook()
                    wb.remove(wb.active)

                    # 每分類一個 sheet，依櫃位排序
                    full_df = pd.DataFrame(cab_rows)
                    for cat_name, cat_group in full_df.groupby("分類", sort=True):
                        ws = wb.create_sheet(title=str(cat_name)[:31])
                        ws.page_setup.orientation = "portrait"
                        ws.page_setup.paperSize = ws.PAPERSIZE_A4
                        ws.page_setup.fitToWidth = 1
                        ws.page_setup.fitToHeight = 0
                        ws.sheet_properties.pageSetUpPr.fitToPage = True

                        headers = ["櫃位", "品項", "廠1", "廠2", "庫存"]
                        widths = [6, 22, 5, 5, 7]
                        for ci, (h, w) in enumerate(zip(headers, widths), 1):
                            cell = ws.cell(row=1, column=ci, value=h)
                            cell.font = Font(bold=True, size=11)
                            cell.border = border_t
                            cell.alignment = Alignment(horizontal="center")
                            ws.column_dimensions[get_column_letter(ci)].width = w

                        sorted_group = cat_group.sort_values(["櫃位", "品項"]).reset_index(drop=True)
                        prev_cab = None
                        row_num = 2
                        for _, r in sorted_group.iterrows():
                            cur_cab = r["櫃位"]
                            # 櫃位切換 → 上一行加雙線
                            if prev_cab is not None and cur_cab != prev_cab and row_num > 2:
                                for ci in range(1, len(headers) + 1):
                                    cell = ws.cell(row=row_num - 1, column=ci)
                                    cell.border = Border(left=thin_s, right=thin_s, top=thin_s, bottom=double_s)

                            show_cab = cur_cab if cur_cab != prev_cab else ""
                            prev_cab = cur_cab

                            vals = [show_cab, r["品項"], r["廠1"], r["廠2"], r["庫存"]]
                            for ci, v in enumerate(vals, 1):
                                cell = ws.cell(row=row_num, column=ci, value=v)
                                cell.font = Font(size=11)
                                cell.border = border_t
                                if ci != 2:
                                    cell.alignment = Alignment(horizontal="center")
                            row_num += 1

                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    st.download_button("📥 下載", data=buf.getvalue(),
                        file_name=f"櫃位分類_{selected_clinic}_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)
            else:
                st.info("尚無品項資料")


# ══════════════════════════════════════════════
#  6. 叫貨出表（per-clinic 廠牌）
# ══════════════════════════════════════════════

def page_order():
    st.header("🛒 叫貨出表")

    selected_clinic = st.session_state.get("selected_clinic", "澤豐")
    if selected_clinic == "合併檢視":
        st.info("請先選擇單一診所。")
        return

    clinic_id = get_clinic_id(selected_clinic)
    if not clinic_id:
        return

    sb = get_supabase_client()

    settings = sb.table("system_settings").select("*").execute().data
    safety = float(next(s["value"] for s in settings if s["key"] == "safety_factor"))
    multiplier = float(next(s["value"] for s in settings if s["key"] == "stock_target_multiplier"))

    products = sb.table("products").select(
        "id, name, category_id, categories(name), units(name)"
    ).execute().data
    products = sort_products_by_bopomofo(products)

    brands = load_brands()
    brand_map = {b["id"]: b["name"] for b in brands}
    brand_names = [b["name"] for b in brands]

    # per-clinic 資料
    cs_data = sb.table("clinic_stock").select(
        "product_id, current_stock, brand1_id, is_active"
    ).eq("clinic_id", clinic_id).execute().data
    cs_map = {s["product_id"]: s for s in cs_data}

    logs = sb.table("inventory_logs").select("product_id, consumed_qty").eq("clinic_id", clinic_id).execute().data
    consumed_data = defaultdict(list)
    for l in logs:
        if l["consumed_qty"] and l["consumed_qty"] > 0:
            consumed_data[l["product_id"]].append(float(l["consumed_qty"]))

    st.subheader("步驟一：編輯叫貨清單")
    show_all = st.checkbox("顯示所有品項", value=False)

    order_rows = []
    for p in products:
        cs = cs_map.get(p["id"])
        if not cs or not cs.get("is_active", True):
            continue

        current = float(cs["current_stock"])
        vals = consumed_data.get(p["id"], [])
        avg = sum(vals) / len(vals) if vals else 0
        auto = current < avg * safety and avg > 0
        suggested = max(0, round(avg * multiplier - current, 1)) if auto else 0

        order_rows.append({
            "勾選": auto, "品項": p["name"], "分類": p["categories"]["name"],
            "單位": p["units"]["name"], "目前庫存": current,
            "建議數量": suggested, "叫貨數量": suggested,
            "廠牌": brand_map.get(cs.get("brand1_id"), ""),
        })

    order_df = pd.DataFrame(order_rows)
    display = order_df if show_all else order_df[order_df["勾選"]]

    if display.empty:
        st.success("所有品項庫存充足！")
        return

    edited = st.data_editor(
        display[["勾選", "品項", "分類", "單位", "目前庫存", "建議數量", "叫貨數量", "廠牌"]],
        use_container_width=True, hide_index=True,
        column_config={
            "勾選": st.column_config.CheckboxColumn("✅"),
            "品項": st.column_config.TextColumn(disabled=True),
            "分類": st.column_config.TextColumn(disabled=True),
            "單位": st.column_config.TextColumn(disabled=True),
            "目前庫存": st.column_config.NumberColumn(disabled=True, format="%.1f"),
            "建議數量": st.column_config.NumberColumn(disabled=True, format="%.1f"),
            "叫貨數量": st.column_config.NumberColumn("數量 ✏️", min_value=0, format="%.1f"),
            "廠牌": st.column_config.SelectboxColumn(options=brand_names),
        },
        height=min(len(display) * 35 + 38, 500), key="order_editor",
    )

    st.divider()
    st.subheader("步驟二：依廠牌分組")

    selected = edited[(edited["勾選"]) & (edited["叫貨數量"] > 0)]
    if selected.empty:
        st.info("請勾選要叫貨的品項")
        return

    brand_dfs = {}
    for brand_name, group in selected.groupby("廠牌"):
        if not brand_name:
            brand_name = "未指定"
        st.markdown(f"### 【{brand_name}】")
        bdf = group[["品項", "叫貨數量", "單位"]].copy()
        edited_brand = st.data_editor(bdf, use_container_width=True, hide_index=True,
                                       column_config={"品項": st.column_config.TextColumn(disabled=True),
                                                      "單位": st.column_config.TextColumn(disabled=True),
                                                      "叫貨數量": st.column_config.NumberColumn(format="%.1f")},
                                       key=f"brand_{brand_name}")
        brand_dfs[brand_name] = edited_brand
        st.caption(f"{len(edited_brand)} 品項")

    st.divider()
    st.subheader("步驟三：匯出")

    col1, col2 = st.columns(2)
    with col1:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for name, bdf in brand_dfs.items():
                bdf.to_excel(writer, sheet_name=name[:31], index=False)
        st.download_button("📥 合併版 .xlsx", data=buf.getvalue(),
                           file_name=f"叫貨單_{selected_clinic}_{date.today()}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, type="primary")
    with col2:
        all_items = pd.concat([bdf.assign(廠牌=n) for n, bdf in brand_dfs.items()], ignore_index=True)
        st.download_button("📥 CSV", data=all_items.to_csv(index=False).encode("utf-8-sig"),
                           file_name=f"叫貨單_{selected_clinic}_{date.today()}.csv",
                           mime="text/csv", use_container_width=True)


# ══════════════════════════════════════════════
#  7. 系統設定（含分類/廠牌/單位的編輯刪除）
# ══════════════════════════════════════════════

def page_settings():
    role = st.session_state.user["role"]
    if role not in ("admin", "manager"):
        st.warning("僅管理者可使用此功能")
        return

    st.header("⚙️ 系統設定")
    sb = get_supabase_client()

    if role == "admin":
        tab_params, tab_cats, tab_brands, tab_units, tab_users = st.tabs([
            "📊 參數", "📂 分類", "🏭 廠牌", "📏 單位", "👤 帳號"
        ])
    else:
        tab_params, tab_cats, tab_brands, tab_units = st.tabs([
            "📊 參數", "📂 分類", "🏭 廠牌", "📏 單位"
        ])
        tab_users = None

    # ── 參數 ──
    with tab_params:
        settings = sb.table("system_settings").select("*").execute().data
        with st.form("settings_form"):
            new_values = {}
            for s in settings:
                new_values[s["key"]] = st.number_input(s["description"], value=float(s["value"]),
                                                        step=0.1, format="%.1f", key=f"s_{s['key']}")
            if st.form_submit_button("儲存", type="primary"):
                for k, v in new_values.items():
                    sb.table("system_settings").update({"value": str(v)}).eq("key", k).execute()
                st.success("已更新")

    # ── 分類 ──
    with tab_cats:
        cats = load_categories()
        units_list = load_units()

        st.subheader("現有分類")
        for c in cats:
            col1, col2, col3 = st.columns([4, 1, 1])
            with col1:
                new_name = st.text_input("名稱", value=c["name"], key=f"cat_name_{c['id']}", label_visibility="collapsed")
            with col2:
                if st.button("💾", key=f"cat_save_{c['id']}"):
                    sb.table("categories").update({"name": new_name}).eq("id", c["id"]).execute()
                    st.success("已更新")
                    load_categories.clear()
            with col3:
                if st.button("🗑️", key=f"cat_del_{c['id']}"):
                    try:
                        sb.table("categories").delete().eq("id", c["id"]).execute()
                        st.success("已刪除")
                        load_categories.clear()
                    except Exception:
                        st.error("無法刪除（有品項使用中）")

        with st.form("add_cat"):
            st.subheader("新增分類")
            nc = st.text_input("名稱", key="new_cat_name")
            unit_names = [u["name"] for u in units_list]
            nu = st.selectbox("預設單位", unit_names, key="new_cat_unit")
            ns = st.text_input("預設規格", key="new_cat_spec")
            if st.form_submit_button("新增"):
                if nc:
                    uid = next(u["id"] for u in units_list if u["name"] == nu)
                    sb.table("categories").insert({"name": nc, "default_unit_id": uid, "default_spec_note": ns or None}).execute()
                    st.success(f"已新增：{nc}")
                    load_categories.clear()

    # ── 廠牌 ──
    with tab_brands:
        blist = load_brands()
        st.subheader("現有廠牌")
        for b in blist:
            col1, col2, col3 = st.columns([4, 1, 1])
            with col1:
                bn = st.text_input("名稱", value=b["name"], key=f"br_{b['id']}", label_visibility="collapsed")
            with col2:
                if st.button("💾", key=f"br_save_{b['id']}"):
                    sb.table("brands").update({"name": bn}).eq("id", b["id"]).execute()
                    st.success("已更新")
                    load_brands.clear()
            with col3:
                if st.button("🗑️", key=f"br_del_{b['id']}"):
                    try:
                        sb.table("brands").delete().eq("id", b["id"]).execute()
                        st.success("已刪除")
                        load_brands.clear()
                    except Exception:
                        st.error("無法刪除（有品項使用中）")

        with st.form("add_brand"):
            nb = st.text_input("新廠牌名稱")
            if st.form_submit_button("新增"):
                if nb:
                    sb.table("brands").insert({"name": nb}).execute()
                    st.success(f"已新增：{nb}")
                    load_brands.clear()

    # ── 單位 ──
    with tab_units:
        ulist = load_units()
        st.subheader("現有單位")
        for u in ulist:
            col1, col2, col3 = st.columns([4, 1, 1])
            with col1:
                un = st.text_input("名稱", value=u["name"], key=f"un_{u['id']}", label_visibility="collapsed")
            with col2:
                if st.button("💾", key=f"un_save_{u['id']}"):
                    sb.table("units").update({"name": un}).eq("id", u["id"]).execute()
                    st.success("已更新")
                    load_units.clear()
            with col3:
                if st.button("🗑️", key=f"un_del_{u['id']}"):
                    try:
                        sb.table("units").delete().eq("id", u["id"]).execute()
                        st.success("已刪除")
                        load_units.clear()
                    except Exception:
                        st.error("無法刪除（有品項使用中）")

        with st.form("add_unit"):
            nu = st.text_input("新單位名稱")
            if st.form_submit_button("新增"):
                if nu:
                    sb.table("units").insert({"name": nu}).execute()
                    st.success(f"已新增：{nu}")
                    load_units.clear()

    # ── 帳號（僅 admin）──
    if tab_users is None:
        return
    with tab_users:
        users = sb.table("users").select("id, username, role, display_name, clinic_id, clinics(name)").execute().data
        clinics_data = sb.table("clinics").select("id, name").execute().data
        clinic_options = {"不限": None}
        for c in clinics_data:
            clinic_options[c["name"]] = c["id"]
        clinic_id_to_name = {c["id"]: c["name"] for c in clinics_data}
        role_options = ["staff", "manager", "admin"]

        st.subheader("現有帳號")
        for u in users:
            with st.expander(f"👤 {u['display_name'] or u['username']}（{u['role']}）"):
                col1, col2 = st.columns(2)
                with col1:
                    new_display = st.text_input("顯示名稱", value=u["display_name"] or "", key=f"ud_{u['id']}")
                    new_role = st.selectbox("角色", role_options,
                                           index=role_options.index(u["role"]), key=f"ur_{u['id']}")
                with col2:
                    current_clinic = clinic_id_to_name.get(u["clinic_id"], "不限")
                    new_clinic = st.selectbox("所屬診所", list(clinic_options.keys()),
                                             index=list(clinic_options.keys()).index(current_clinic)
                                             if current_clinic in clinic_options else 0,
                                             key=f"uc_{u['id']}")
                    new_pw = st.text_input("重設密碼（留空不改）", type="password", key=f"up_{u['id']}")

                col_save, col_del = st.columns([3, 1])
                with col_save:
                    if st.button("💾 儲存", key=f"us_{u['id']}"):
                        updates = {
                            "display_name": new_display or u["username"],
                            "role": new_role,
                            "clinic_id": clinic_options[new_clinic],
                        }
                        if new_pw:
                            from auth import hash_password
                            updates["password_hash"] = hash_password(new_pw)
                        sb.table("users").update(updates).eq("id", u["id"]).execute()
                        st.success("已更新")
                with col_del:
                    if u["username"] != "admin":
                        if st.button("🗑️ 刪除", key=f"udel_{u['id']}"):
                            sb.table("users").delete().eq("id", u["id"]).execute()
                            st.success("已刪除")
                            st.rerun()
                    else:
                        st.caption("(不可刪除)")

        st.divider()
        with st.form("add_user"):
            st.subheader("新增帳號")
            col1, col2 = st.columns(2)
            with col1:
                nu = st.text_input("帳號 *")
                np = st.text_input("密碼 *", type="password")
            with col2:
                nd = st.text_input("顯示名稱")
                nr = st.selectbox("角色", role_options)
            nc = st.selectbox("所屬診所", list(clinic_options.keys()))

            if st.form_submit_button("新增帳號", type="primary"):
                if not nu or not np:
                    st.error("帳號和密碼為必填")
                else:
                    from auth import hash_password
                    try:
                        sb.table("users").insert({
                            "username": nu, "password_hash": hash_password(np),
                            "display_name": nd or nu, "role": nr, "clinic_id": clinic_options[nc],
                        }).execute()
                        st.success(f"已新增：{nu}")
                    except Exception as e:
                        st.error(f"失敗：{e}")
